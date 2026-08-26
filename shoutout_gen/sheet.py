"""Read the shout-out spreadsheet (CSV or XLSX) into ordered messages.

The sheet is expected to be a Google-Form style export: a timestamp column and
a free-text shout-out column, header names unknown in advance. Columns are
detected by header keywords first and by content second, and both can be
forced from the CLI when detection guesses wrong.

Ordering rule: by timestamp ascending (stable), which becomes click order on
the slide. Rows without a parseable timestamp keep their sheet position after
the timestamped ones. Blank messages are dropped; duplicates are kept because
"bao" submitted by three people is three shout-outs.
"""
from __future__ import annotations

import csv
import re
from dataclasses import dataclass
from datetime import datetime
from pathlib import Path

from openpyxl import load_workbook

TIME_HEADER_RE = re.compile(r"time|date|submitted", re.I)
TEXT_HEADER_RE = re.compile(r"shout|message|text|response", re.I)

# Formats Google Forms / Sheets / Excel exports actually produce.
_TIMESTAMP_FORMATS = (
    "%m/%d/%Y %H:%M:%S",
    "%m/%d/%Y %H:%M",
    "%m/%d/%y %H:%M:%S",
    "%m/%d/%y %H:%M",
    "%Y-%m-%d %H:%M:%S",
    "%Y-%m-%d %H:%M",
    "%m/%d/%Y",
    "%Y-%m-%d",
)


@dataclass(frozen=True)
class Shoutout:
    """One message to put on the slide; ``row`` is the 1-based sheet row for error messages."""

    text: str
    timestamp: datetime | None
    row: int


class SheetFormatError(ValueError):
    """The sheet has no usable message column (or the forced column doesn't exist)."""


# --------------------------------------------------------------------------- #
# Public interface
# --------------------------------------------------------------------------- #
def read_shoutouts(path: Path, text_column: str | None = None, time_column: str | None = None) -> list[Shoutout]:
    """Load, clean and order the shout-outs in ``path`` (.csv or .xlsx)."""
    header, rows = _load_table(path)
    if not header:
        raise SheetFormatError(f"{path} is empty")
    time_col = _pick_time_column(header, rows, time_column)
    text_col = _pick_text_column(header, rows, text_column, exclude=time_col)

    shoutouts: list[Shoutout] = []
    for row_no, row in enumerate(rows, start=2):  # row 1 is the header
        text = _clean_text(row.get(text_col))
        if not text:
            continue
        stamp = parse_timestamp(row.get(time_col)) if time_col else None
        shoutouts.append(Shoutout(text, stamp, row_no))
    # Stable sort: timestamped rows ascending, then untimestamped in sheet order.
    shoutouts.sort(key=lambda s: (s.timestamp is None, s.timestamp or datetime.min))
    return shoutouts


def parse_timestamp(value) -> datetime | None:
    """datetime from a cell that may already be a datetime, an ISO string, or a US-style string."""
    if isinstance(value, datetime):
        return value
    if value is None:
        return None
    text = str(value).strip()
    if not text:
        return None
    try:
        return datetime.fromisoformat(text)
    except ValueError:
        pass
    for fmt in _TIMESTAMP_FORMATS:
        try:
            return datetime.strptime(text, fmt)
        except ValueError:
            continue
    return None


# --------------------------------------------------------------------------- #
# Private mechanics
# --------------------------------------------------------------------------- #
def _load_table(path: Path) -> tuple[list[str], list[dict[str, object]]]:
    """Header names and one dict per data row, for CSV or the first sheet of an XLSX."""
    suffix = path.suffix.lower()
    if suffix == ".csv":
        with path.open(encoding="utf-8-sig", newline="") as fh:
            reader = csv.DictReader(fh)
            header = list(reader.fieldnames or [])
            return header, list(reader)
    if suffix in (".xlsx", ".xlsm"):
        ws = load_workbook(path, read_only=True, data_only=True).worksheets[0]
        rows_iter = ws.iter_rows(values_only=True)
        header_cells = next(rows_iter, None) or ()
        header = [str(c).strip() if c is not None else "" for c in header_cells]
        rows = [dict(zip(header, r)) for r in rows_iter if any(c not in (None, "") for c in r)]
        return header, rows
    raise SheetFormatError(f"Unsupported spreadsheet type {suffix!r}; use .csv or .xlsx")


def _pick_time_column(header: list[str], rows: list[dict], forced: str | None) -> str | None:
    if forced is not None:
        if forced not in header:
            raise SheetFormatError(f"time column {forced!r} not in header {header}")
        return forced
    for name in header:
        if TIME_HEADER_RE.search(name):
            return name
    # No keyword match: take a column whose values mostly parse as timestamps.
    for name in header:
        values = [r.get(name) for r in rows if r.get(name) not in (None, "")]
        if values and sum(parse_timestamp(v) is not None for v in values) >= 0.8 * len(values):
            return name
    return None


def _pick_text_column(header: list[str], rows: list[dict], forced: str | None, exclude: str | None) -> str:
    if forced is not None:
        if forced not in header:
            raise SheetFormatError(f"text column {forced!r} not in header {header}")
        return forced
    candidates = [h for h in header if h != exclude and h]
    if not candidates:
        raise SheetFormatError(f"no message column found in header {header}")
    for name in candidates:
        if TEXT_HEADER_RE.search(name):
            return name
    # Fall back to the wordiest column: shout-outs are the longest free text on a form.
    def mean_len(name: str) -> float:
        lengths = [len(str(r.get(name) or "")) for r in rows]
        return sum(lengths) / len(lengths) if lengths else 0.0

    return max(candidates, key=mean_len)


def _clean_text(value) -> str:
    """Normalise line endings and trim; the slide renders the text verbatim otherwise."""
    if value is None:
        return ""
    return str(value).replace("\r\n", "\n").replace("\r", "\n").strip()
